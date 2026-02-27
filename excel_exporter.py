"""
excel_exporter.py
─────────────────
Exports a BuildingReport to a professional 3-sheet Excel workbook:
  Sheet 1 — Room Schedule     (every room, GFA / NOFA breakdown)
  Sheet 2 — Area Summary      (floor-by-floor and building totals)
  Sheet 3 — Efficiency Ratios (NOFA/GFA, benchmarks, warnings)

Usage:
    from excel_exporter import export_to_excel
    export_to_excel(report, "area_schedule.xlsx", project_name="Tower A")
"""

from __future__ import annotations
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from area_calculator import BuildingReport, RoomResult


# ─── Style helpers ────────────────────────────────────────────────────────────

NAVY   = "1F3864"
BLUE   = "2E5FA3"
LBLUE  = "D9E1F2"
YELLOW = "FFF2CC"
GREEN  = "E2EFDA"
RED_BG = "FCE4D6"
WHITE  = "FFFFFF"
LGREY  = "F2F2F2"
WARN   = "FF0000"

def _font(size=10, bold=False, color="000000", italic=False):
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)

def _fill(hex_col):
    return PatternFill("solid", fgColor=hex_col, start_color=hex_col)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(color="BFBFBF"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _num_format(ws, row, col, fmt):
    ws.cell(row=row, column=col).number_format = fmt

def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _title_row(ws, row, ncols, text, bg=NAVY, font_size=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font    = _font(font_size, bold=True, color="FFFFFF")
    c.fill    = _fill(bg)
    c.alignment = _align("center", "center")
    ws.row_dimensions[row].height = 24

def _section_row(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font    = _font(10, bold=True, color="FFFFFF")
    c.fill    = _fill(BLUE)
    c.alignment = _align("left", "center")
    ws.row_dimensions[row].height = 18

def _header_row(ws, row, headers, bg=BLUE):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = _font(9, bold=True, color="FFFFFF")
        c.fill      = _fill(bg)
        c.alignment = _align("center", "center", wrap=True)
        c.border    = _border()
    ws.row_dimensions[row].height = 28

def _data_cell(ws, row, col, value, bg=WHITE, bold=False,
               align="left", border=True, italic=False, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(9, bold=bold, color=color, italic=italic)
    c.fill      = _fill(bg)
    c.alignment = _align(align, "center")
    if border: c.border = _border()
    return c


# ─── Sheet 1: Room Schedule ───────────────────────────────────────────────────

def _write_room_schedule(ws, report: BuildingReport, project_name: str, date_str: str):
    _set_widths(ws, [6, 28, 10, 12, 12, 12, 12, 12, 14, 18, 30])
    NC = 11

    _title_row(ws, 1, NC, f"Room Schedule — {project_name}")
    _title_row(ws, 2, NC,
        f"Building Type: {report.building_type.value.title()}   |   "
        f"Date: {date_str}   |   Spec: PNAP APP-2 & APP-151 (Rev. Jul 2025) — QS reviewed Feb 2026",
        bg=BLUE, font_size=9)

    headers = [
        "#", "Room / Space", "Floor",
        "Polygon\nArea (m²)", "GFA\nRule",
        "GFA\nContrib (m²)", "NOFA\nRule",
        "NOFA\nContrib (m²)",
        "Saleable\nArea (m²)",   # NEW — Q4.3 confirmed by QS
        "APP-151\nConcession", "Notes"
    ]
    _header_row(ws, 3, headers)

    # Saleable Area header note
    ws.cell(row=3, column=9).font = _font(9, bold=True, color="FFFFFF")
    ws.cell(row=3, column=9).fill = _fill("5B4A8A")   # distinct purple column

    # Group results by floor
    floors: dict[str, list[RoomResult]] = {}
    for r in report.rooms:
        floors.setdefault(r.input.floor, []).append(r)

    row = 4
    overall_num = 1
    for floor_label, rooms in floors.items():
        _section_row(ws, row, NC, f"Floor: {floor_label}")
        row += 1

        floor_polygon = floor_gfa = floor_nofa = floor_saleable = 0.0

        for i, r in enumerate(rooms):
            bg = WHITE if i % 2 == 0 else LGREY
            c  = r.classification

            has_warn = "⚠️" in (c.gfa_note + c.nofa_note)
            if has_warn: bg = YELLOW

            # Saleable Area = GFA contribution of habitable rooms only
            # (Cap. 621 definition — includes all areas inside unit except
            #  common areas, plant, shafts. Pending full AP/QS definition.)
            is_habitable = c.nofa_rule.value == "full"
            saleable = r.gfa_area_m2 if is_habitable else 0.0

            _data_cell(ws, row, 1,  overall_num,         bg, align="center")
            _data_cell(ws, row, 2,  r.input.label,        bg)
            _data_cell(ws, row, 3,  r.input.floor,        bg, align="center")
            _data_cell(ws, row, 4,  r.area_m2,            bg, align="right")
            _data_cell(ws, row, 5,  c.gfa_rule.value,     bg, align="center",
                       color="006400" if c.gfa_rule.value == "full"
                       else ("CC0000" if c.gfa_rule.value == "excluded" else "8B4513"))
            _data_cell(ws, row, 6,  r.gfa_area_m2,        bg, align="right")
            _data_cell(ws, row, 7,  c.nofa_rule.value,    bg, align="center",
                       color="006400" if c.nofa_rule.value == "full" else "CC0000")
            _data_cell(ws, row, 8,  r.nofa_area_m2,       bg, align="right")

            # Saleable Area cell — distinct styling
            sa_cell = ws.cell(row=row, column=9, value=saleable if saleable > 0 else "—")
            sa_cell.font      = _font(9, color="3B2A6E" if saleable > 0 else "AAAAAA", italic=(saleable == 0))
            sa_cell.fill      = _fill("F0EDFA")
            sa_cell.border    = _border()
            sa_cell.alignment = _align("right", "center")
            if saleable > 0:
                sa_cell.number_format = '#,##0.00'

            _data_cell(ws, row, 10, c.concession_item,    bg, align="center",
                       italic=bool(c.concession_item))
            note = c.gfa_note if "⚠️" in c.gfa_note else (c.nofa_note if "⚠️" in c.nofa_note else "")
            _data_cell(ws, row, 11, note, bg, color="CC0000" if note else "000000",
                       italic=bool(note))

            for col in [4, 6, 8]:
                ws.cell(row=row, column=col).number_format = '#,##0.00'

            floor_polygon  += r.area_m2
            floor_gfa      += r.gfa_area_m2
            floor_nofa     += r.nofa_area_m2
            floor_saleable += saleable
            overall_num    += 1
            row            += 1

        # Floor subtotal
        for col in range(1, NC+1):
            c2 = ws.cell(row=row, column=col)
            c2.fill   = _fill(LBLUE)
            c2.border = _border("9E9E9E")
        ws.cell(row=row, column=2, value=f"Floor Subtotal — {floor_label}").font = _font(9, bold=True)
        ws.cell(row=row, column=2).fill = _fill(LBLUE)
        ws.cell(row=row, column=2).alignment = _align()
        ws.cell(row=row, column=2).border = _border()
        for col, val in [(4, floor_polygon), (6, floor_gfa), (8, floor_nofa), (9, floor_saleable)]:
            c2 = ws.cell(row=row, column=col, value=val)
            c2.font   = _font(9, bold=True)
            c2.fill   = _fill(LBLUE) if col != 9 else _fill("DDD5F0")
            c2.border = _border()
            c2.alignment = _align("right", "center")
            c2.number_format = '#,##0.00'
        row += 2

    # Grand total
    _section_row(ws, row, NC, "Grand Total")
    row += 1
    for col in range(1, NC+1):
        c2 = ws.cell(row=row, column=col)
        c2.fill   = _fill(NAVY)
        c2.border = _border()

    # Calculate total saleable from rooms
    total_saleable = sum(
        r.gfa_area_m2 for r in report.rooms
        if r.classification.nofa_rule.value == "full"
    )

    for col, val, label in [
        (2,  "BUILDING TOTAL",    None),
        (4,  report.total_polygon_m2, "Polygon m²"),
        (6,  report.total_gfa_m2,     "GFA m²"),
        (8,  report.total_nofa_m2,    "NOFA m²"),
        (9,  total_saleable,          "Saleable m²"),
    ]:
        c2 = ws.cell(row=row, column=col, value=val)
        c2.font      = _font(10, bold=True, color="FFFFFF")
        c2.fill      = _fill(NAVY)
        c2.border    = _border()
        c2.alignment = _align("right" if isinstance(val, float) else "left", "center")
        if isinstance(val, float):
            c2.number_format = '#,##0.00'

    # Saleable area disclaimer
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
    disc = ws.cell(row=row, column=1,
        value="⚠️  Saleable Area (Cap. 621): currently approximated as habitable GFA areas. "
              "Pending full AP/QS definition — do not use for sales documentation without verification.")
    disc.font      = _font(9, italic=True, color="8B4513")
    disc.fill      = _fill(YELLOW)
    disc.border    = _border("8B4513")
    disc.alignment = _align("left", "center", wrap=True)
    ws.row_dimensions[row].height = 28


# ─── Sheet 2: Area Summary ────────────────────────────────────────────────────

def _write_area_summary(ws, report: BuildingReport, project_name: str, date_str: str):
    _set_widths(ws, [30, 18, 18, 18, 18, 28])
    NC = 6

    _title_row(ws, 1, NC, f"Area Summary — {project_name}")
    _title_row(ws, 2, NC,
        f"Building Type: {report.building_type.value.title()}   |   Date: {date_str}",
        bg=BLUE, font_size=9)

    # ── Floor-by-floor breakdown ─────────────────────────────────────────────
    _section_row(ws, 3, NC, "Floor-by-Floor Breakdown")
    _header_row(ws, 4, ["Floor", "Polygon Area (m²)", "GFA (m²)",
                         "NOFA (m²)", "NOFA/GFA (%)", "Notes"])

    floors: dict[str, dict] = {}
    for r in report.rooms:
        f = r.input.floor
        if f not in floors:
            floors[f] = {"polygon": 0.0, "gfa": 0.0, "nofa": 0.0}
        floors[f]["polygon"] += r.area_m2
        floors[f]["gfa"]     += r.gfa_area_m2
        floors[f]["nofa"]    += r.nofa_area_m2

    row = 5
    for i, (floor_label, totals) in enumerate(floors.items()):
        bg = WHITE if i % 2 == 0 else LGREY
        ratio = totals["nofa"] / totals["gfa"] if totals["gfa"] > 0 else 0
        _data_cell(ws, row, 1, floor_label,       bg, bold=True)
        _data_cell(ws, row, 2, totals["polygon"], bg, align="right")
        _data_cell(ws, row, 3, totals["gfa"],     bg, align="right")
        _data_cell(ws, row, 4, totals["nofa"],    bg, align="right")
        _data_cell(ws, row, 5, ratio,             bg, align="right")
        _data_cell(ws, row, 6, "",                bg)
        for col in [2, 3, 4]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        ws.cell(row=row, column=5).number_format = '0.0%'
        row += 1

    # Building total row
    bg = LBLUE
    ratio_total = report.total_nofa_m2 / report.total_gfa_m2 if report.total_gfa_m2 > 0 else 0
    _data_cell(ws, row, 1, "BUILDING TOTAL", bg, bold=True)
    _data_cell(ws, row, 2, report.total_polygon_m2, bg, bold=True, align="right")
    _data_cell(ws, row, 3, report.total_gfa_m2,     bg, bold=True, align="right")
    _data_cell(ws, row, 4, report.total_nofa_m2,    bg, bold=True, align="right")
    _data_cell(ws, row, 5, ratio_total,             bg, bold=True, align="right")
    _data_cell(ws, row, 6, "",                      bg)
    for col in [2, 3, 4]:
        ws.cell(row=row, column=col).number_format = '#,##0.00'
    ws.cell(row=row, column=5).number_format = '0.0%'
    row += 2

    # ── APP-151 Concessions ──────────────────────────────────────────────────
    _section_row(ws, row, NC, "APP-151 GFA Concessions")
    row += 1
    _header_row(ws, row, ["Concession Item", "Total Polygon (m²)",
                           "Effective GFA (m²)", "Subject to 10% Cap",
                           "BEAM Plus Req'd", "Status"])
    row += 1

    for i, con in enumerate(report.concessions):
        bg = RED_BG if con.cap_warning else (WHITE if i % 2 == 0 else LGREY)
        _data_cell(ws, row, 1, con.item,                    bg)
        _data_cell(ws, row, 2, con.total_area_m2,           bg, align="right")
        _data_cell(ws, row, 3, con.effective_gfa_m2,        bg, align="right")
        _data_cell(ws, row, 4, "YES" if con.subject_to_cap else "No",
                   bg, align="center",
                   color="CC0000" if con.subject_to_cap else "006400")
        _data_cell(ws, row, 5, "YES" if con.requires_beam_plus else "No",
                   bg, align="center",
                   color="CC0000" if con.requires_beam_plus else "006400")
        status = "⚠️ CAP EXCEEDED" if con.cap_warning else (
                  "BD Approval Required" if con.requires_beam_plus else "Confirmed Exempt")
        _data_cell(ws, row, 6, status, bg,
                   color="CC0000" if "CAP" in status else
                   ("8B4513" if "BD" in status else "006400"))
        for col in [2, 3]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        row += 1

    row += 1
    _data_cell(ws, row, 1, "Capped Concessions Total",     LBLUE, bold=True)
    _data_cell(ws, row, 2, report.capped_total_m2,         LBLUE, bold=True, align="right")
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    _data_cell(ws, row, 3, report.cap_limit_m2,            LBLUE, bold=True, align="right")
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    _data_cell(ws, row, 4, f"{report.cap_utilisation_pct:.1f}% of 10% cap used",
               LBLUE, bold=True)
    row += 1

    if report.cap_exceeded:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
        c2 = ws.cell(row=row, column=1,
                     value="⚠️ APP-151 10% CAP EXCEEDED — BD approval required before submission.")
        c2.font      = _font(10, bold=True, color="CC0000")
        c2.fill      = _fill(RED_BG)
        c2.alignment = _align("center", "center")
        c2.border    = _border("CC0000")


# ─── Sheet 3: Efficiency Ratios ───────────────────────────────────────────────

_BENCHMARKS = {
    "residential":  {"nofa_gfa_low": 0.65, "nofa_gfa_high": 0.80},
    "non_domestic": {"nofa_gfa_low": 0.55, "nofa_gfa_high": 0.70},
    "composite":    {"nofa_gfa_low": 0.60, "nofa_gfa_high": 0.75},
    "hotel":        {"nofa_gfa_low": 0.55, "nofa_gfa_high": 0.70},
}

def _write_efficiency_ratios(ws, report: BuildingReport, project_name: str, date_str: str):
    _set_widths(ws, [32, 20, 20, 20, 30])
    NC = 5

    _title_row(ws, 1, NC, f"Efficiency Ratios — {project_name}")
    _title_row(ws, 2, NC,
        f"Building Type: {report.building_type.value.title()}   |   Date: {date_str}",
        bg=BLUE, font_size=9)

    bm = _BENCHMARKS.get(report.building_type.value, _BENCHMARKS["residential"])
    nofa_gfa = report.nofa_gfa_ratio

    # ── Key ratios table ─────────────────────────────────────────────────────
    _section_row(ws, 3, NC, "Key Area Metrics")
    _header_row(ws, 4, ["Metric", "Value", "Unit", "HK Benchmark", "Status"])

    metrics = [
        ("Total GFA",         report.total_gfa_m2,   "m²", "—"),
        ("Total NOFA",        report.total_nofa_m2,  "m²", "—"),
        ("Total Polygon Area",report.total_polygon_m2,"m²","—"),
        ("NOFA / GFA Ratio",  nofa_gfa,              "%",
         f"{bm['nofa_gfa_low']*100:.0f}% – {bm['nofa_gfa_high']*100:.0f}%"),
        ("10% Cap Utilisation",report.cap_utilisation_pct / 100, "%", "< 100%"),
    ]

    for i, (metric, value, unit, benchmark) in enumerate(metrics, 5):
        bg = WHITE if i % 2 == 0 else LGREY

        # Colour-code status
        if metric == "NOFA / GFA Ratio":
            in_range = bm["nofa_gfa_low"] <= value <= bm["nofa_gfa_high"]
            status   = "✅ Within benchmark" if in_range else "⚠️ Outside benchmark"
            s_color  = "006400" if in_range else "CC0000"
            bg_s     = GREEN if in_range else RED_BG
        elif metric == "10% Cap Utilisation":
            status   = "✅ Within cap" if not report.cap_exceeded else "❌ Cap exceeded"
            s_color  = "006400" if not report.cap_exceeded else "CC0000"
            bg_s     = GREEN if not report.cap_exceeded else RED_BG
        else:
            status, s_color, bg_s = "—", "000000", bg

        _data_cell(ws, i, 1, metric,    bg)
        c2 = _data_cell(ws, i, 2, value, bg, align="right")
        if unit == "%":
            c2.number_format = "0.0%"
        else:
            c2.number_format = "#,##0.00"
        _data_cell(ws, i, 3, unit,      bg, align="center")
        _data_cell(ws, i, 4, benchmark, bg, align="center")
        _data_cell(ws, i, 5, status,    bg_s, color=s_color)

    # ── Benchmark reference table ────────────────────────────────────────────
    row = 11
    _section_row(ws, row, NC, "HK Market Benchmarks (Reference)")
    row += 1
    _header_row(ws, row, ["Building Type", "NOFA/GFA Low", "NOFA/GFA High",
                           "Source", "Notes"])
    row += 1
    for btype, bvals in _BENCHMARKS.items():
        bg = WHITE if row % 2 == 0 else LGREY
        _data_cell(ws, row, 1, btype.replace("_", " ").title(), bg)
        c_lo = _data_cell(ws, row, 2, bvals["nofa_gfa_low"],  bg, align="right")
        c_hi = _data_cell(ws, row, 3, bvals["nofa_gfa_high"], bg, align="right")
        c_lo.number_format = "0%"
        c_hi.number_format = "0%"
        _data_cell(ws, row, 4, "HK industry convention", bg, italic=True)
        _data_cell(ws, row, 5, "Pending AP/QS confirmation", bg, italic=True,
                   color="8B4513")
        row += 1

    # ── Warnings ─────────────────────────────────────────────────────────────
    if report.warnings:
        row += 1
        _section_row(ws, row, NC, "Warnings")
        row += 1
        for w in report.warnings:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=NC)
            c2 = ws.cell(row=row, column=1, value=f"⚠️  {w}")
            c2.font      = _font(9, color="CC0000")
            c2.fill      = _fill(RED_BG)
            c2.border    = _border("CC0000")
            c2.alignment = _align("left", "center", wrap=True)
            ws.row_dimensions[row].height = 30
            row += 1


# ─── Public API ───────────────────────────────────────────────────────────────

def export_to_excel(
    report:       BuildingReport,
    output_path:  str,
    project_name: str = "Floor Plan Area Calculator",
) -> str:
    """
    Export a BuildingReport to a 3-sheet Excel workbook.

    Args:
        report:       BuildingReport from AreaCalculator.calculate()
        output_path:  Output .xlsx file path.
        project_name: Project name shown in headers.

    Returns:
        Absolute path to the saved file.
    """
    wb = Workbook()
    date_str = datetime.today().strftime("%d %b %Y")

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Room Schedule"
    _write_room_schedule(ws1, report, project_name, date_str)

    # Sheet 2
    ws2 = wb.create_sheet("Area Summary")
    _write_area_summary(ws2, report, project_name, date_str)

    # Sheet 3
    ws3 = wb.create_sheet("Efficiency Ratios")
    _write_efficiency_ratios(ws3, report, project_name, date_str)

    # Freeze header rows
    for ws in [ws1, ws2, ws3]:
        ws.freeze_panes = "A4"
        ws.sheet_view.showGridLines = False

    wb.save(output_path)
    return str(output_path)
